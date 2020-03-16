#!/usr/bin/env python
#-*- coding:utf-8 -*-
# author:Administrator
# datetime:2020/2/27 12:19
# software: PyCharm

from tinydb import TinyDB
from tinydb import Query
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment


class Product( object):
    __prod_id = 0

    def __init__(self, product, amount, start_date, end_date, price, ):
        self.product = product
        self.amount = amount
        self.start_date = start_date
        self.end_date = end_date
        self.price = price
        self.payback_dict={}
        self.id = self.__id_code()

    @property
    def day_interest(self):
        day_interest = self.amount * self.price /365
        return day_interest

    @property
    def month_interest(self):
        month_interest = self.amount * self.price /12
        return month_interest

    @property
    def status(self):
        all_payback = 0
        for k,v in self.payback_dict.items():
            all_payback += v
        if self.amount > all_payback:
            return '未结清'
        elif self.amount == all_payback:
            return '已结清'
        else:
            return '输入数据有误'

    @classmethod
    def __id_code(cls):
        cls.__prod_id += 1
        num = str(cls.__prod_id)
        return str('产品' + num )

    def payback(self,payback_amount, payback_date):
        self.amount -= payback_amount
        self.payback_dict[payback_date] = payback_amount

    def info(self):
        info_dict = {
            '用信品种': self.product,
            '用信敞口': self.amount,
            '起始日期': self.start_date,
            '到期日期': self.end_date,
            '价格': self.price,
            '还款列表': self.payback_dict,
            '形态': self.status ,
        }
        return info_dict

class Project(object):
    __proj_id = 0

    def __init__(self, name, credit_amount, manager):
        self.name = name
        self.credit_amount = credit_amount
        self.manager = manager
        self.status = '营销中'
        self.credit_used = 0
        self.credit_balance = self.credit_amount - self.credit_used
        self.product_list = []
        self.id = self.__id_code()

    @property
    def product_count(self):
        return len(self.product_list)

    @classmethod
    def __id_code(cls):
        cls.__proj_id += 1
        num = str(cls.__proj_id)
        return str('项目' + num)

    def __check_same(self,product_obj):
        flag = False
        for product_exsited in self.product_list:
            if product_exsited.id == product_obj.id:
                flag = True
                break
        return flag

    def stat_refresh(self, new_stat):
        if new_stat in ['营销中','调查中','评议会','已上报','贷审会','已否决']:
            self.status = new_stat
            self.credit_balance = float(0)
        elif new_stat == '已批未投':
            self.status = new_stat
            self.credit_balance = self.credit_amount
        elif new_stat == '已投放':
            self.status = new_stat
            self.credit_balance = self.credit_amount - self.credit_used
        else:
            pass

    def add_prod(self, product_obj):
        if self.__check_same(product_obj):
            pass
        else:
            if self.credit_balance >= product_obj.amount:
                self.credit_used += product_obj.amount
                self.stat_refresh('已投放')
                self.product_list.append(product_obj)
            else:
                pass

    def min_prod(self, product_obj):
        if product_obj in self.product_list:
            self.product_list.remove(product_obj)
            self.credit_used -= product_obj.amount
        else:
            print('无此笔贷款，请仔细核对！')



    def info(self):
        info_dict={
            '企业名称':self.name,
            '客户经理':self.manager,
            '营销状态':self.status,
            '授信额度':self.credit_amount,
            '可用额度':self.credit_balance,
            '已用额度':self.credit_used,
            '用信笔数':self.product_count,
        }
        return info_dict


class ProjPool(object):

    def __init__(self):
        self.db = TinyDB('pp.json')
        self.table_proj = self.db.table('proj')
        self.table_prod = self.db.table('prod')
        self.proj_list = []

    def __check_same(self, obj):
        flag = False
        if len(self.proj_list) == 0:
            return flag
        else:
            for proj in self.proj_list:
                if proj.name == obj.name and \
                        proj.id == obj.id:
                    flag = True
                    break
            return flag

    def add_project(self, proj_obj):
        if self.__check_same(proj_obj):
            pass
        else:
            self.proj_list.append(proj_obj)
            self.table_proj.insert(proj_obj.info())
            if len(proj_obj.product_list) > 0:
                for prod_obj in proj_obj.product_list:
                    prod_dict = prod_obj.info()
                    prod_dict.update({
                        '企业名称': proj_obj.name,
                        '客户经理': proj_obj.manager,
                    })
                    self.table_prod.insert(prod_dict)

    def __ws_proj(self,ws_proj):
        Q = Query()
        title_proj = [
            '企业名称','客户经理','授信额度',
            '已用额度','可用额度','营销状态','用信笔数'
        ]
        for title in title_proj:
            ws_proj.cell(row=1,
                         column=title_proj.index(title)+1,
                         value=title)
            for docs in self.table_proj:
                ws_proj.cell(row=docs.doc_id+1,
                             column= title_proj.index(title)+1,
                             value = docs[title])

    def __ws_prod(self,ws_prod):
        title_prod = [
            '企业名称','客户经理','用信品种','用信敞口',
            '起始日期','到期日期','价格', '形态'
        ]
        for title in title_prod:
            ws_prod.cell(row=1,
                         column=title_prod.index(title)+1,
                         value=title)
            for docs in self.table_prod:
                ws_prod.cell(row=docs.doc_id+1,
                             column= title_prod.index(title)+1,
                             value = docs[title])

    def output(self):
        wb = Workbook()
        ws_proj = wb.active
        ws_proj.title = '项目进度表'
        ws_prod = wb.create_sheet('投放明细表')
        ws_proj.sheet_properties.tabColor = '1072BA'
        ws_prod.sheet_properties.tabColor = '43b2bc'
        self.__ws_proj(ws_proj)
        self.__ws_prod(ws_prod)
        wb.save('项目进度表.xlsx')


